from __future__ import annotations

import streamlit as st
import streamlit.components.v1 as components
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse, unquote
import time
import io
import re
import json
import xml.etree.ElementTree as ET
import zlib
import base64
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# openpyxl rejects control chars in cells (common in scraped HTML)
_EXCEL_CTRL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
SHARE_PACK_VERSION = 1
# Browser / proxy limits on query string length — above this, use JSON file only
SHARE_URL_MAX_CHARS = 3200

# ─────────────────────────────────────────────
# Page config
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="UX Architect Pro",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
# Categorie emergenti dall'analisi (URL + menu), colori assegnati a posteriori
# ─────────────────────────────────────────────
# Pastelli distinti; testo nel diagramma e sui badge via _contrast_label_hex (nero o bianco).
CATEGORY_COLOR_PALETTE = [
    "#B8D4EC", "#B5E8D1", "#DDD6FE", "#FBCFE8", "#F5E0A8", "#FED7AA",
    "#BFDBFE", "#A7F3D0", "#FECDD3", "#C4B5FD", "#99F6E4", "#E9D5FF",
    "#D9F99D", "#FBC4A2", "#A5D8FF", "#CCD5E0",
]
# Grigio‑lavanda chiaro se la categoria non è in mappa.
CATEGORY_COLOR_FALLBACK = "#E2E8F0"
# Fallback quando non si ricava una sezione dal sito (UI / export in italiano)
CAT_FALLBACK = "Altro"


def category_palette_map(categories: list[str]) -> dict[str, str]:
    """Un colore stabile per ogni nome categoria (ordinamento alfabetico = ordine palette)."""
    uniq = sorted(set(categories))
    return {
        c: CATEGORY_COLOR_PALETTE[i % len(CATEGORY_COLOR_PALETTE)]
        for i, c in enumerate(uniq)
    }


def _contrast_label_hex(hex_color: str) -> str:
    """#111827 su sfondi chiari, #ffffff su sfondi scuri (diagramma, badge)."""
    h = (hex_color or "").strip().lstrip("#")
    if len(h) != 6:
        return "#111827"
    try:
        r = int(h[0:2], 16)
        g = int(h[2:4], 16)
        b = int(h[4:6], 16)
    except ValueError:
        return "#111827"
    lum = (0.299 * r + 0.587 * g + 0.114 * b) / 255
    return "#111827" if lum > 0.62 else "#ffffff"


# ─────────────────────────────────────────────
# Styling
# ─────────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

    html, body, [class*="css"] {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    }

    .block-container { max-width: 1200px; padding-top: 1.25rem; }

    .sidebar-app-title {
        font-size: 0.72rem;
        font-weight: 800;
        letter-spacing: 0.14em;
        color: #111827;
        margin: 0 0 1.1rem 0;
        line-height: 1.35;
        padding-bottom: 0.85rem;
        border-bottom: 1px solid #eaedf0;
    }

    .stat-card {
        background: #fff;
        border: 1px solid #eaedf0;
        border-radius: 12px;
        padding: 1.4rem 1.5rem;
        height: 100%;
        transition: box-shadow 0.2s ease;
    }
    .stat-card:hover { box-shadow: 0 4px 20px rgba(0,0,0,0.06); }
    .stat-card h3 {
        margin: 0;
        font-size: 0.75rem;
        font-weight: 600;
        color: #6b7280;
        text-transform: uppercase;
        letter-spacing: 0.6px;
    }
    .stat-card .value {
        font-size: 1.85rem;
        font-weight: 700;
        margin: 0.4rem 0 0;
        color: #111827;
    }

    .log-entry {
        font-family: 'SF Mono', 'Fira Code', 'JetBrains Mono', monospace;
        font-size: 0.78rem;
        padding: 4px 0;
        border-bottom: 1px solid #f3f4f6;
        color: #374151;
        line-height: 1.6;
    }
    .log-ok  { color: #047857; }
    .log-err { color: #b91c1c; }
    .log-info { color: #1d4ed8; }

    .category-badge {
        display: inline-block;
        padding: 4px 12px;
        border-radius: 20px;
        font-size: 0.75rem;
        font-weight: 600;
        letter-spacing: 0.2px;
    }

    .section-title {
        font-size: 1.15rem;
        font-weight: 700;
        color: #111827;
        margin-bottom: 0.3rem;
        letter-spacing: -0.3px;
    }
    .section-subtitle {
        font-size: 0.88rem;
        color: #6b7280;
        margin-bottom: 1.2rem;
    }

    div[data-testid="stSidebar"] {
        background: #fafbfc;
        border-right: 1px solid #eaedf0;
    }
    div[data-testid="stSidebar"] .stMarkdown h3 {
        font-size: 0.85rem;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        color: #6b7280;
    }

    .legend-item {
        display: flex;
        align-items: center;
        gap: 8px;
        padding: 3px 0;
        font-size: 0.85rem;
        color: #374151;
    }
    .legend-dot {
        width: 10px;
        height: 10px;
        border-radius: 50%;
        flex-shrink: 0;
    }

    div[data-testid="stTabs"] button {
        font-weight: 600;
        font-size: 0.85rem;
        letter-spacing: 0.1px;
    }

    .stExpander {
        border: 1px solid #eaedf0 !important;
        border-radius: 10px !important;
        margin-bottom: 6px;
    }

    .empty-state {
        text-align: center;
        padding: 5rem 2rem;
    }
    .empty-state h2 {
        color: #111827;
        font-weight: 700;
        font-size: 1.5rem;
        margin-bottom: 0.5rem;
    }
    .empty-state p {
        color: #6b7280;
        font-size: 1rem;
        max-width: 480px;
        margin: 0 auto;
        line-height: 1.6;
    }

    .sidebar-source-hint {
        font-size: 0.8rem;
        color: #6b7280;
        line-height: 1.45;
        margin: 0 0 0.75rem 0;
    }
    .sidebar-panel-label {
        font-size: 0.72rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.45px;
        color: #6b7280;
        margin: 0 0 0.35rem 0;
    }

    /* CTA nero; accenti neutri ardesia (slider, link) */
    :root {
        --ux-cta: #111827;
        --ux-cta-hover: #27272a;
        --ux-accent: #475569;
        --ux-accent-hover: #334155;
        --ux-accent-soft: rgba(71, 85, 105, 0.12);
    }

    div[data-testid="stSidebar"] button[kind="primary"],
    section[data-testid="stSidebar"] button[kind="primary"],
    .stButton > button[kind="primary"] {
        background-color: var(--ux-cta) !important;
        border-color: var(--ux-cta) !important;
        color: #ffffff !important;
        font-weight: 600 !important;
    }
    div[data-testid="stSidebar"] button[kind="primary"]:hover,
    section[data-testid="stSidebar"] button[kind="primary"]:hover,
    .stButton > button[kind="primary"]:hover {
        background-color: var(--ux-cta-hover) !important;
        border-color: var(--ux-cta-hover) !important;
    }

    /* Slider fill & radio selected */
    div[data-testid="stSlider"] [role="slider"] {
        background-color: var(--ux-accent) !important;
    }
    div[data-baseweb="slider"] [data-testid="stThumbValue"] {
        color: var(--ux-accent);
    }

    label[data-testid="stWidgetLabel"] a {
        color: var(--ux-accent);
    }

    /* Hero vuoto — solo testo centrato (nessuna illustrazione) */
    .ux-empty-hero {
        padding: 2rem 1rem 3.5rem;
        margin: 0 auto;
    }
    .ux-empty-stack {
        max-width: 520px;
        margin: 0 auto;
        display: flex;
        flex-direction: column;
        align-items: center;
    }
    .ux-empty-title {
        color: #111827;
        font-weight: 800;
        font-size: 1.55rem;
        letter-spacing: -0.35px;
        margin: 0 0 0.4rem 0;
        line-height: 1.4;
        text-align: center;
        width: 100%;
        max-width: 520px;
    }
    .ux-wave-hand {
        display: inline-block;
        animation: uxHandWave 2.2s ease-in-out infinite;
        transform-origin: 72% 85%;
    }
    @keyframes uxHandWave {
        0%, 100% { transform: rotate(0deg); }
        20% { transform: rotate(16deg); }
        40% { transform: rotate(-10deg); }
        60% { transform: rotate(14deg); }
        80% { transform: rotate(-6deg); }
    }
    @media (prefers-reduced-motion: reduce) {
        .ux-wave-hand { animation: none; }
    }
    .ux-empty-subtitle {
        margin: 0 0 1rem 0;
        padding: 0;
        font-size: 1.12rem;
        font-weight: 700;
        letter-spacing: -0.2px;
        color: #374151;
        line-height: 1.35;
        text-align: center;
        width: 100%;
        max-width: 520px;
    }
    .ux-empty-hint {
        color: #4b5563;
        font-size: 1.02rem;
        line-height: 1.7;
        margin: 0;
        width: 100%;
        max-width: 520px;
        text-align: center;
        box-sizing: border-box;
    }
</style>
""", unsafe_allow_html=True)

# ═════════════════════════════════════════════
# CORE FUNCTIONS
# ═════════════════════════════════════════════

def normalize_url(url: str) -> str:
    """Strip fragment and trailing slash for deduplication."""
    parsed = urlparse(url)
    path = parsed.path.rstrip("/") or "/"
    return f"{parsed.scheme}://{parsed.netloc}{path}"


def is_same_domain(url: str, base_domain: str) -> bool:
    parsed = urlparse(url)
    return parsed.netloc == base_domain or parsed.netloc == ""


def _collect_links(tags, current_url: str, base_domain: str) -> list[str]:
    """Resolve a list of <a> tags to absolute same-domain URLs."""
    out: list[str] = []
    seen: set[str] = set()
    for tag in tags:
        href = (tag.get("href") or "").strip()
        if not href or href.startswith(("#", "mailto:", "tel:", "javascript:")):
            continue
        absolute = normalize_url(urljoin(current_url, href))
        if absolute not in seen and is_same_domain(absolute, base_domain):
            seen.add(absolute)
            out.append(absolute)
    return out


def extract_nav_links(soup: BeautifulSoup, current_url: str, base_domain: str) -> list[str]:
    """Extract links from main navigation elements (header, nav, menus)."""
    nav_tags = []

    for nav in soup.find_all("nav"):
        nav_tags.extend(nav.find_all("a", href=True))

    header = soup.find("header")
    if header:
        nav_tags.extend(header.find_all("a", href=True))

    for el in soup.find_all(class_=re.compile(
        r"(main-menu|primary-menu|nav-menu|site-nav|navbar|menu-item|"
        r"main-navigation|primary-navigation|mega-menu)", re.I
    )):
        nav_tags.extend(el.find_all("a", href=True))

    for el in soup.find_all(id=re.compile(
        r"(menu|nav|navigation|main-menu|primary-menu)", re.I
    )):
        nav_tags.extend(el.find_all("a", href=True))

    return _collect_links(nav_tags, current_url, base_domain)


def extract_all_links(soup: BeautifulSoup, current_url: str, base_domain: str) -> list[str]:
    """Extract all same-domain links from the page."""
    return _collect_links(soup.find_all("a", href=True), current_url, base_domain)


def extract_navigations(soup: BeautifulSoup, current_url: str,
                        base_domain: str) -> dict[str, list[dict]]:
    """Extract all navigation structures from the page.

    Returns a dict mapping professional labels to their menu trees:
        {"Main Navigation (Header)": [...], "Secondary Navigation (Footer)": [...], ...}
    Each tree item: {"label": str, "url": str, "children": [...]}
    """
    def _resolve(href: str) -> str:
        if not href or href.startswith(("#", "mailto:", "tel:", "javascript:")):
            return ""
        u = normalize_url(urljoin(current_url, href))
        return u if is_same_domain(u, base_domain) else ""

    def _is_column_wrapper(li_el) -> bool:
        classes = " ".join(li_el.get("class", [])).lower()
        is_col_class = any(k in classes for k in (
            "column", "mega-col", "menu-col",
            "sub-menu-column", "mega-menu-column",
        ))
        if not is_col_class:
            return False
        for child in li_el.children:
            if hasattr(child, "name") and child.name in ("a", "span", "strong", "b"):
                if child.get_text(strip=True):
                    return False
        direct_text = li_el.find(string=True, recursive=False)
        if direct_text and direct_text.strip():
            return False
        return True

    def _parse_ul(ul_el) -> list[dict]:
        items: list[dict] = []
        for li in ul_el.find_all("li", recursive=False):
            if _is_column_wrapper(li):
                sub_ul = li.find("ul")
                if sub_ul:
                    items.extend(_parse_ul(sub_ul))
                continue
            a = li.find("a", recursive=False) or li.find("a")
            label = ""
            url = ""
            if a:
                label = a.get_text(strip=True)
                url = _resolve(a.get("href", ""))
            if not label:
                span = li.find(["span", "strong", "b"], recursive=False)
                if span:
                    label = span.get_text(strip=True)
            if not label:
                continue
            children: list[dict] = []
            sub_ul = li.find("ul")
            if sub_ul:
                children = _parse_ul(sub_ul)
            items.append({"label": label, "url": url, "children": children})
        return items

    def _parse_nav_element(nav_el) -> list[dict]:
        top_ul = nav_el.find("ul")
        if top_ul:
            return _parse_ul(top_ul)
        items: list[dict] = []
        for a in nav_el.find_all("a", href=True):
            label = a.get_text(strip=True)
            url = _resolve(a.get("href", ""))
            if label:
                items.append({"label": label, "url": url, "children": []})
        return items

    navigations: dict[str, list[dict]] = {}
    used_navs: set[int] = set()

    # ── Main Navigation (Header) ──
    header = soup.find("header")
    if header:
        header_navs = header.find_all("nav")
        if header_navs:
            main_nav = max(header_navs, key=lambda n: len(n.find_all("a")))
            items = _parse_nav_element(main_nav)
            if items:
                navigations["Main Navigation (Header)"] = items
                used_navs.add(id(main_nav))
            for hn in header_navs:
                if id(hn) != id(main_nav) and id(hn) not in used_navs:
                    hi = _parse_nav_element(hn)
                    if hi:
                        navigations["Utility Navigation (Header)"] = hi
                        used_navs.add(id(hn))
                        break

    if "Main Navigation (Header)" not in navigations:
        for candidate in soup.find_all("nav"):
            if id(candidate) in used_navs:
                continue
            role = (candidate.get("aria-label") or candidate.get("role") or "").lower()
            classes = " ".join(candidate.get("class", [])).lower()
            if any(k in classes for k in ("main", "primary", "site-nav", "navbar")) or \
               any(k in role for k in ("main", "primary", "navigation")):
                items = _parse_nav_element(candidate)
                if items:
                    navigations["Main Navigation (Header)"] = items
                    used_navs.add(id(candidate))
                    break

    if "Main Navigation (Header)" not in navigations:
        all_navs = [n for n in soup.find_all("nav") if id(n) not in used_navs]
        if all_navs:
            best = max(all_navs, key=lambda n: len(n.find_all("a")))
            items = _parse_nav_element(best)
            if items:
                navigations["Main Navigation (Header)"] = items
                used_navs.add(id(best))

    # ── Secondary Navigation (Footer) ──
    footer = soup.find("footer")
    if footer:
        footer_navs = footer.find_all("nav")
        if footer_navs:
            for fn in footer_navs:
                if id(fn) not in used_navs:
                    items = _parse_nav_element(fn)
                    if items:
                        navigations["Secondary Navigation (Footer)"] = items
                        used_navs.add(id(fn))
                        break
        if "Secondary Navigation (Footer)" not in navigations:
            footer_uls = footer.find_all("ul", recursive=True)
            all_footer_items: list[dict] = []
            for ul in footer_uls:
                if ul.find_parent("nav") and id(ul.find_parent("nav")) in used_navs:
                    continue
                all_footer_items.extend(_parse_ul(ul))
            if all_footer_items:
                navigations["Secondary Navigation (Footer)"] = all_footer_items

    # ── Sidebar Navigation ──
    sidebar_tags = soup.find_all(["aside", "div"], class_=re.compile(
        r"(sidebar|side-nav|side-menu|lateral)", re.I
    ))
    for sb in sidebar_tags:
        sb_nav = sb.find("nav")
        el = sb_nav if sb_nav else sb
        if id(el) in used_navs:
            continue
        items = _parse_nav_element(el) if el.name == "nav" else []
        if not items:
            sb_ul = el.find("ul")
            if sb_ul:
                items = _parse_ul(sb_ul)
        if items:
            navigations["Sidebar Navigation"] = items
            used_navs.add(id(el))
            break

    return navigations


def flatten_nav_urls(navigations: dict[str, list[dict]]) -> list[str]:
    """Collect all URLs from all navigation sections."""
    def _flatten(items: list[dict]) -> list[str]:
        urls: list[str] = []
        for item in items:
            if item["url"]:
                urls.append(item["url"])
            urls.extend(_flatten(item["children"]))
        return urls

    all_urls: list[str] = []
    for items in navigations.values():
        all_urls.extend(_flatten(items))
    return all_urls


def _path_key_from_url(url: str) -> str:
    parsed = urlparse(url)
    parts = [x for x in parsed.path.strip("/").split("/") if x]
    return "/".join(parts) if parts else "(home)"


SM_EXCEL_MAX_LEVELS = 12


def build_sitemap_tree_excel_rows(
    results: list[dict],
    navigations: dict[str, list[dict]] | None,
) -> list[dict]:
    """Righe come il tab Sitemap: zona, livelli gerarchici (etichette), anteprima ASCII, URL."""
    rows: list[dict] = []
    if not results:
        return rows
    url_to_page = {p["url"]: p for p in results}

    def walk_menu(items: list[dict], prefix: str, zone: str, ancestors: list[str]) -> None:
        for i, item in enumerate(items):
            last = i == len(items) - 1
            connector = "└── " if last else "├── "
            page = url_to_page.get(item["url"]) if item.get("url") else None
            label = item.get("label") or ""
            if page and page.get("title"):
                label = page["title"]
            tree_line = f"{prefix}{connector}{label}"
            levels = ancestors + [label]
            rows.append({
                "navigazione": zone,
                "levels": levels,
                "albero": tree_line,
                "url": item.get("url") or "",
            })
            extension = "    " if last else "│   "
            if item.get("children"):
                walk_menu(item["children"], prefix + extension, zone, ancestors + [label])

    if navigations:
        home_url = normalize_url(
            results[0]["url"].split(urlparse(results[0]["url"]).path)[0] + "/"
        )
        home_page = url_to_page.get(home_url)
        root_title = (home_page.get("title") if home_page else None) or urlparse(results[0]["url"]).netloc
        rows.append({
            "navigazione": "(root)",
            "levels": [root_title],
            "albero": root_title,
            "url": home_url,
        })

        for nav_name, nav_items in navigations.items():
            rows.append({
                "navigazione": nav_name,
                "levels": [nav_name],
                "albero": f"── {nav_name} ──",
                "url": "",
            })
            walk_menu(nav_items, "", nav_name, [])

        all_nav_urls = set(flatten_nav_urls(navigations))
        remaining = [
            p for p in results
            if p["url"] not in all_nav_urls
            and p.get("status_code", 200) != 404
            and p["url"] != home_url
        ]
        if remaining:
            rows.append({
                "navigazione": "Altre pagine",
                "levels": ["Altre pagine"],
                "albero": f"── Altre pagine ({len(remaining)}) ──",
                "url": "",
            })
            for j, page in enumerate(remaining):
                last = j == len(remaining) - 1
                conn = "└── " if last else "├── "
                lab = page.get("title") or page["url"]
                rows.append({
                    "navigazione": "Altre pagine",
                    "levels": ["Altre pagine", lab],
                    "albero": f"{conn}{lab}",
                    "url": page["url"],
                })
    else:
        path_to_title: dict[str, str] = {}
        for page in results:
            parsed = urlparse(page["url"])
            path = parsed.path.strip("/") or "(home)"
            title = (page.get("title") or "").strip()
            if title:
                path_to_title[path] = title

        tree: dict = {}
        for page in results:
            parsed = urlparse(page["url"])
            parts = [p for p in parsed.path.strip("/").split("/") if p]
            if not parts:
                parts = ["(home)"]
            node = tree
            for part in parts:
                if part not in node:
                    node[part] = {}
                node = node[part]

        def walk_url_tree(
            node: dict, prefix: str, current_path: str, ancestors: list[str],
        ) -> None:
            items = sorted(node.keys())
            for i, key in enumerate(items):
                last = i == len(items) - 1
                connector = "└── " if last else "├── "
                full_path = f"{current_path}/{key}".strip("/") if current_path else key
                label = path_to_title.get(full_path, key)
                url_match = ""
                for p in results:
                    if _path_key_from_url(p["url"]) == full_path:
                        url_match = p["url"]
                        break
                levels = ancestors + [label]
                rows.append({
                    "navigazione": "Struttura URL",
                    "levels": levels,
                    "albero": f"{prefix}{connector}{label}",
                    "url": url_match,
                })
                extension = "    " if last else "│   "
                walk_url_tree(node[key], prefix + extension, full_path, ancestors + [label])

        root_label = urlparse(results[0]["url"]).netloc
        root_title = path_to_title.get("(home)", root_label)
        rows.append({
            "navigazione": "Struttura URL",
            "levels": [root_title],
            "albero": root_title,
            "url": "",
        })
        walk_url_tree(tree, "", "", [root_title])

    return rows


def _excel_url_display(url: str, max_len: int = 58) -> str:
    u = (url or "").strip()
    if len(u) <= max_len:
        return u
    return u[: max_len - 1] + "…"


def mermaid_to_png_bytes(mermaid_code: str) -> bytes | None:
    """Render Mermaid in PNG via Kroki (POST: affidabile anche per diagrammi grandi)."""
    code = (mermaid_code or "").strip()
    if not code:
        return None
    for _post in (
        lambda: requests.post(
            "https://kroki.io/mermaid/png",
            data=code.encode("utf-8"),
            headers={"Content-Type": "text/plain; charset=utf-8"},
            timeout=60,
        ),
        lambda: requests.post(
            "https://kroki.io/mermaid/png",
            json={"diagram_source": code},
            headers={"Content-Type": "application/json"},
            timeout=60,
        ),
    ):
        try:
            r = _post()
            if r.status_code == 200 and r.content.startswith(b"\x89PNG\r\n\x1a\n"):
                return r.content
        except Exception:
            pass
    return None


def _xml_local_tag(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def _parse_sitemap_xml(xml_text: str) -> tuple[list[str], list[str]]:
    """Estrae (URL pagina, URL sitemap figlia) da sitemap XML — senza lxml (Cloud-friendly)."""
    page_urls: list[str] = []
    child_sitemaps: list[str] = []
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return page_urls, child_sitemaps

    if _xml_local_tag(root.tag) == "sitemapindex":
        for sm in root:
            if _xml_local_tag(sm.tag) != "sitemap":
                continue
            for ch in sm:
                if _xml_local_tag(ch.tag) == "loc" and ch.text and ch.text.strip():
                    child_sitemaps.append(ch.text.strip())
    else:
        for uel in root:
            if _xml_local_tag(uel.tag) != "url":
                continue
            for ch in uel:
                if _xml_local_tag(ch.tag) == "loc" and ch.text and ch.text.strip():
                    page_urls.append(ch.text.strip())
    return page_urls, child_sitemaps


def fetch_sitemap_urls(base_url: str, session: requests.Session, base_domain: str) -> list[str]:
    """Try to fetch URLs from sitemap.xml or robots.txt sitemap reference."""
    urls: list[str] = []
    sitemap_locations = [
        f"{base_url.rstrip('/')}/sitemap.xml",
        f"{base_url.rstrip('/')}/sitemap_index.xml",
    ]

    try:
        robots_resp = session.get(f"{base_url.rstrip('/')}/robots.txt", timeout=8)
        if robots_resp.status_code == 200:
            for line in robots_resp.text.splitlines():
                if line.strip().lower().startswith("sitemap:"):
                    sm_url = line.split(":", 1)[1].strip()
                    if sm_url and sm_url not in sitemap_locations:
                        sitemap_locations.insert(0, sm_url)
    except Exception:
        pass

    def collect_from_document(xml_text: str) -> None:
        page_locs, child_locs = _parse_sitemap_xml(xml_text)
        for raw in page_locs:
            u = normalize_url(raw)
            if is_same_domain(u, base_domain):
                urls.append(u)
        # Un livello di sitemap figlie (come la versione BeautifulSoup/lxml)
        for child_url in child_locs:
            try:
                sub_resp = session.get(child_url, timeout=8)
                if sub_resp.status_code != 200:
                    continue
                p2, _ = _parse_sitemap_xml(sub_resp.text)
                for raw in p2:
                    u = normalize_url(raw)
                    if is_same_domain(u, base_domain):
                        urls.append(u)
            except Exception:
                pass

    for sm_url in sitemap_locations:
        try:
            resp = session.get(sm_url, timeout=8)
            if resp.status_code != 200:
                continue
            collect_from_document(resp.text)
            if urls:
                break
        except Exception:
            continue

    return list(dict.fromkeys(urls))


def extract_page_data(url: str, response: requests.Response, soup: BeautifulSoup) -> dict:
    """Extract structural data from a page."""
    title_tag = soup.find("title")
    title = title_tag.get_text(strip=True) if title_tag else ""

    meta_desc_tag = soup.find("meta", attrs={"name": "description"})
    meta_desc = meta_desc_tag.get("content", "") if meta_desc_tag else ""

    og_type_tag = soup.find("meta", attrs={"property": "og:type"})
    og_type = og_type_tag.get("content", "") if og_type_tag else ""

    h1_tag = soup.find("h1")
    h1 = h1_tag.get_text(strip=True) if h1_tag else ""

    h2_tags = soup.find_all("h2")
    h2_list = [h.get_text(strip=True) for h in h2_tags]

    body = soup.find("body")
    text = body.get_text(" ", strip=True) if body else soup.get_text(" ", strip=True)
    word_count = len(text.split())

    breadcrumbs = extract_breadcrumbs(soup)

    return {
        "url": url,
        "status_code": response.status_code,
        "title": title,
        "meta_description": meta_desc,
        "og_type": og_type,
        "h1": h1,
        "h2_list": h2_list,
        "word_count": word_count,
        "breadcrumbs": breadcrumbs,
    }


def extract_breadcrumbs(soup: BeautifulSoup) -> str:
    """Try to extract breadcrumbs from structured data or common patterns."""
    ld_scripts = soup.find_all("script", type="application/ld+json")
    for script in ld_scripts:
        try:
            import json
            data = json.loads(script.string or "")
            items = None
            if isinstance(data, dict) and data.get("@type") == "BreadcrumbList":
                items = data.get("itemListElement", [])
            elif isinstance(data, list):
                for d in data:
                    if isinstance(d, dict) and d.get("@type") == "BreadcrumbList":
                        items = d.get("itemListElement", [])
                        break
            if items:
                names = [i.get("item", {}).get("name", i.get("name", "")) for i in sorted(items, key=lambda x: x.get("position", 0))]
                return " > ".join(n for n in names if n)
        except Exception:
            pass

    bc_nav = soup.find("nav", attrs={"aria-label": re.compile(r"breadcrumb", re.I)})
    if not bc_nav:
        bc_nav = soup.find(class_=re.compile(r"breadcrumb", re.I))
    if bc_nav:
        links = bc_nav.find_all(["a", "span", "li"])
        parts = [l.get_text(strip=True) for l in links if l.get_text(strip=True)]
        if parts:
            return " > ".join(parts)

    return ""


# ─────────────────────────────────────────────
# Categorizzazione emergente (menu + struttura URL del sito analizzato)
# ─────────────────────────────────────────────

_LOCALE_SEGMENTS = frozenset({
    "en", "it", "fr", "de", "es", "pt", "nl", "pl", "ru", "zh", "ja", "ko",
    "ar", "sv", "da", "no", "fi", "cs", "sk", "hu", "ro", "bg", "el", "tr", "uk",
})
_SKIP_PATH_SEGMENTS = frozenset({
    "wp", "wp-content", "static", "assets", "cdn", "uploads", "media", "files",
    "public", "dist", "js", "css", "img", "images", "fonts", "vendors",
})
_META_PATH_SEGMENTS = frozenset({
    "category", "tag", "author", "page", "archives", "taxonomy", "tipo", "categoria",
})


def _humanize_url_segment(seg: str) -> str:
    s = unquote(seg).strip().replace("-", " ").replace("_", " ")
    parts = [w.capitalize() for w in s.split() if w]
    label = " ".join(parts)[:50].strip()
    return label or CAT_FALLBACK


def _category_from_url_path(url: str, base_url: str) -> str | None:
    """Primo segmento di percorso significativo (dopo eventuale lingua WP / cartelle tecniche)."""
    parsed = urlparse(url)
    base_parsed = urlparse(base_url)
    parts = [p for p in parsed.path.strip("/").split("/") if p]
    if not parts:
        return None
    i = 0
    if len(parts[0]) == 2 and parts[0].lower() in _LOCALE_SEGMENTS:
        i = 1
    if i >= len(parts):
        return None
    seg = parts[i]
    low = seg.lower()
    if low in _SKIP_PATH_SEGMENTS:
        i += 1
        if i >= len(parts):
            return None
        seg = parts[i]
        low = seg.lower()
    if low in _META_PATH_SEGMENTS and i + 1 < len(parts):
        return _humanize_url_segment(parts[i + 1])
    # Homepage del sito come path unico slug
    if len(parts) == 1 and normalize_url(url) == normalize_url(base_url):
        return None
    return _humanize_url_segment(seg)


def _nav_section_label_for_url(
    navigations: dict[str, list[dict]],
    page_url: str,
) -> str | None:
    """Etichetta di sezione dal menu: genitore della voce che punta a questa URL (o la voce stessa)."""
    target = normalize_url(page_url)
    best_depth = -1
    best: str | None = None

    def walk(items: list[dict], ancestors: list[str]) -> None:
        nonlocal best_depth, best
        for item in items:
            lab = (item.get("label") or "").strip()
            children = item.get("children") or []
            chain = ancestors + ([lab] if lab else [])
            u = (item.get("url") or "").strip()
            if u:
                try:
                    if normalize_url(u) == target:
                        d = len(chain)
                        if d > best_depth:
                            best_depth = d
                            if len(chain) >= 2:
                                best = chain[-2]
                            elif chain:
                                best = chain[-1]
                except Exception:
                    pass
            walk(children, chain)

    for items in navigations.values():
        walk(items, [])
    if not best:
        return None
    return " ".join(best.split())[:50] or None


def infer_page_category(
    page_data: dict,
    base_url: str,
    navigations: dict[str, list[dict]] | None = None,
) -> str:
    """Assegna una categoria che emerge dal sito: Home, voce di menu (sezione), o segmento URL."""
    url = page_data["url"]
    parsed = urlparse(url)
    path = unquote(parsed.path).lower()

    nu_page = normalize_url(url)
    nu_base = normalize_url(base_url)
    if nu_page == nu_base:
        return "Home"
    if path in ("", "/", "/index.html", "/index.php", "/home", "/homepage"):
        if urlparse(base_url).path.rstrip("/") in ("", path.rstrip("/")):
            return "Home"

    if navigations:
        nav_lab = _nav_section_label_for_url(navigations, url)
        if nav_lab:
            return nav_lab

    path_cat = _category_from_url_path(url, base_url)
    if path_cat:
        return path_cat

    return CAT_FALLBACK


# ─────────────────────────────────────────────
# Crawler
# ─────────────────────────────────────────────

CRAWL_WORKERS = 6
# (connect, read) in seconds — many production sites (WordPress, shared hosting)
# need >10s before the first byte; read timeout applies to the full body download.
CRAWL_HTTP_TIMEOUT = (15, 60)
CRAWL_FETCH_RETRIES = 2  # extra attempts after Timeout / ConnectionError


def crawl_site(start_url: str, max_depth: int, max_pages: int,
               progress_bar, log_container, status_text):
    """Priority crawler with concurrent fetching.

    Uses a ThreadPoolExecutor to fetch pages in batches of CRAWL_WORKERS,
    drastically reducing wall-clock time compared to sequential requests.
    """
    from collections import deque
    from concurrent.futures import ThreadPoolExecutor, as_completed

    parsed_start = urlparse(start_url)
    base_domain = parsed_start.netloc
    start_url = normalize_url(start_url)

    visited: set[str] = set()
    priority_queue: deque[tuple[str, int]] = deque()
    secondary_queue: deque[tuple[str, int]] = deque()
    results: list[dict] = []
    errors_404: list[str] = []
    log_lines: list[str] = []
    nav_urls: set[str] = set()

    session = requests.Session()
    session.headers.update({
        "User-Agent": "UXArchitectPro/1.0 (Streamlit Crawler)",
        "Accept": "text/html,application/xhtml+xml",
        "Accept-Language": "it-IT,it;q=0.9,en;q=0.8",
    })
    adapter = requests.adapters.HTTPAdapter(
        pool_connections=CRAWL_WORKERS,
        pool_maxsize=CRAWL_WORKERS * 2,
    )
    session.mount("https://", adapter)
    session.mount("http://", adapter)

    def add_log(msg: str, level: str = "info"):
        icons = {"info": "🔍", "ok": "✅", "err": "🔴", "warn": "⚠️"}
        css = {"info": "log-info", "ok": "log-ok", "err": "log-err", "warn": "log-err"}
        log_lines.append(f'<div class="log-entry {css.get(level, "")}">{icons.get(level, "")} {msg}</div>')
        log_container.markdown("".join(log_lines[-60:]), unsafe_allow_html=True)

    def update_ui():
        pct = min(len(results) / max_pages, 1.0)
        progress_bar.progress(pct, text=f"Scansionate {len(results)}/{max_pages} pagine")
        q_total = len(priority_queue) + len(secondary_queue)
        status_text.caption(
            f"Pagine visitate: {len(visited)} | In coda: {q_total} | "
            f"Errori 404: {len(errors_404)}"
        )

    def _fetch(url: str):
        """Network I/O — runs in worker threads."""
        last_exc: Exception | None = None
        for attempt in range(1 + CRAWL_FETCH_RETRIES):
            try:
                return session.get(
                    url, timeout=CRAWL_HTTP_TIMEOUT, allow_redirects=True,
                )
            except (requests.Timeout, requests.ConnectionError) as exc:
                last_exc = exc
                if attempt < CRAWL_FETCH_RETRIES:
                    time.sleep(0.6 * (attempt + 1))
        assert last_exc is not None
        raise last_exc

    def _process_response(url: str, depth: int, resp) -> BeautifulSoup | None:
        """Parse response and record page data. Runs in main thread."""
        if resp.status_code == 404:
            errors_404.append(url)
            add_log(f"<b>404</b>: {url}", "err")
            results.append({
                "url": url, "status_code": 404, "title": "", "meta_description": "",
                "og_type": "", "h1": "", "h2_list": [], "word_count": 0,
                "breadcrumbs": "", "_html": "", "depth": depth,
            })
            update_ui()
            return None

        if resp.status_code >= 400:
            add_log(f"HTTP {resp.status_code}: {url}", "warn")
            return None

        content_type = resp.headers.get("Content-Type", "")
        if "text/html" not in content_type:
            return None

        soup = BeautifulSoup(resp.text, "html.parser")
        page_data = extract_page_data(url, resp, soup)
        page_data["_html"] = resp.text[:30_000]
        page_data["depth"] = depth
        page_data["is_nav"] = url in nav_urls
        results.append(page_data)

        title_preview = page_data["title"][:60] or "(no title)"
        tag = " [menu]" if url in nav_urls else ""
        add_log(f"OK — <b>{title_preview}</b>{tag}", "ok")
        update_ui()
        return soup

    def _drain_queue(queue: deque, add_to_secondary: bool = False):
        """Fetch pages from queue concurrently in batches."""
        with ThreadPoolExecutor(max_workers=CRAWL_WORKERS) as pool:
            while queue and len(results) < max_pages:
                batch: list[tuple[str, int]] = []
                while queue and len(batch) < CRAWL_WORKERS:
                    url, depth = queue.popleft()
                    if url in visited or depth > max_depth or len(results) >= max_pages:
                        continue
                    visited.add(url)
                    batch.append((url, depth))

                if not batch:
                    break

                future_map = {
                    pool.submit(_fetch, url): (url, depth)
                    for url, depth in batch
                }

                for future in as_completed(future_map):
                    if len(results) >= max_pages:
                        break
                    url, depth = future_map[future]
                    try:
                        resp = future.result()
                    except requests.RequestException as exc:
                        add_log(f"Errore di rete: {url} — {exc}", "err")
                        continue
                    except Exception as exc:
                        add_log(f"Errore: {url} — {exc}", "err")
                        continue

                    soup = _process_response(url, depth, resp)
                    if soup and depth < max_depth:
                        new_links = extract_all_links(soup, url, base_domain)
                        if add_to_secondary:
                            for link in new_links:
                                if link not in visited:
                                    secondary_queue.append((link, depth + 1))
                        else:
                            for link in new_links:
                                if link not in visited:
                                    queue.append((link, depth + 1))

    def _process_single(url: str, depth: int) -> BeautifulSoup | None:
        """Fetch and process one page synchronously (used for homepage)."""
        if url in visited or len(results) >= max_pages:
            return None
        visited.add(url)
        try:
            add_log(f"[depth {depth}] {url}")
            resp = _fetch(url)
            return _process_response(url, depth, resp)
        except requests.RequestException as exc:
            add_log(f"Errore di rete: {url} — {exc}", "err")
        except Exception as exc:
            add_log(f"Errore: {url} — {exc}", "err")
        return None

    navigations: dict[str, list[dict]] = {}

    # ── Phase 1: Homepage + extract navigation hierarchies ──
    add_log(f"Avvio crawl di <b>{start_url}</b>")
    add_log("Fase 1 — Analisi homepage e strutture di navigazione", "info")

    home_soup = _process_single(start_url, 0)

    if home_soup:
        navigations = extract_navigations(home_soup, start_url, base_domain)

        if navigations:
            def _log_menu(items: list[dict], indent: int = 0):
                for item in items:
                    prefix = "&nbsp;&nbsp;" * indent
                    add_log(f"{prefix}• <b>{item['label']}</b>", "ok")
                    if item["children"]:
                        _log_menu(item["children"], indent + 1)

            for nav_name, nav_items in navigations.items():
                add_log(f"<b>{nav_name}</b>: {len(nav_items)} voci top-level", "ok")
                _log_menu(nav_items)

            all_nav_urls = flatten_nav_urls(navigations)
            nav_urls.update(all_nav_urls)
            for nav_link in all_nav_urls:
                if nav_link not in visited:
                    priority_queue.append((nav_link, 1))
        else:
            homepage_nav = extract_nav_links(home_soup, start_url, base_domain)
            nav_urls.update(homepage_nav)
            if homepage_nav:
                add_log(f"Trovate <b>{len(homepage_nav)}</b> voci di navigazione (flat)", "ok")
                for nav_link in homepage_nav:
                    if nav_link not in visited:
                        priority_queue.append((nav_link, 1))
            else:
                add_log("Nessun menu trovato, uso tutti i link", "warn")
                for link in extract_all_links(home_soup, start_url, base_domain):
                    if link not in visited:
                        priority_queue.append((link, 1))

    # ── Phase 2: Sitemap.xml ──
    add_log("Fase 2 — Ricerca sitemap.xml", "info")
    sitemap_urls = fetch_sitemap_urls(start_url, session, base_domain)
    if sitemap_urls:
        add_log(f"Sitemap trovata: <b>{len(sitemap_urls)}</b> URL", "ok")
        for sm_url in sitemap_urls:
            if sm_url not in visited and sm_url not in nav_urls:
                secondary_queue.append((sm_url, 1))
    else:
        add_log("Nessuna sitemap trovata", "warn")

    # ── Phase 3: Crawl priority queue (nav links) — concurrent ──
    add_log("Fase 3 — Scansione voci di navigazione", "info")
    _drain_queue(priority_queue, add_to_secondary=True)

    # ── Phase 4: Crawl secondary queue — concurrent ──
    if len(results) < max_pages and secondary_queue:
        add_log("Fase 4 — Scansione pagine secondarie e sitemap", "info")
    _drain_queue(secondary_queue, add_to_secondary=False)

    add_log(f"Crawl completato — <b>{len(results)}</b> pagine analizzate", "ok")
    progress_bar.progress(1.0, text="Crawl completato!")
    return results, errors_404, navigations


# ─────────────────────────────────────────────
# Mermaid diagram generation
# ─────────────────────────────────────────────

def _safe_id(url: str) -> str:
    """Create a valid Mermaid node id from a URL."""
    parsed = urlparse(url)
    path = parsed.path.strip("/") or "home"
    node_id = re.sub(r"[^a-zA-Z0-9]", "_", path)
    if node_id[0].isdigit():
        node_id = "n" + node_id
    return node_id[:60]


def _build_page_tree(results: list[dict]):
    """Build a URL-path tree and a path-to-page lookup from crawl results."""
    path_to_page: dict[str, dict] = {}
    for page in results:
        parsed = urlparse(page["url"])
        path = parsed.path.strip("/") or "(home)"
        path_to_page[path] = page

    tree: dict = {}
    for page in results:
        parsed = urlparse(page["url"])
        parts = [p for p in parsed.path.strip("/").split("/") if p]
        if not parts:
            parts = ["(home)"]
        node = tree
        for part in parts:
            if part not in node:
                node[part] = {}
            node = node[part]

    return tree, path_to_page


def _make_mermaid_id_factory():
    """Return a function that generates unique Mermaid node IDs."""
    used: set[str] = set()

    def make_id(hint: str) -> str:
        nid = re.sub(r"[^a-zA-Z0-9]", "_", hint)[:50]
        if not nid or nid[0].isdigit():
            nid = "n" + (nid or "x")
        orig = nid
        c = 0
        while nid in used:
            c += 1
            nid = f"{orig}_{c}"
        used.add(nid)
        return nid

    return make_id


def _url_to_page(results: list[dict]) -> dict[str, dict]:
    """Map normalized URL -> page data."""
    return {page["url"]: page for page in results}


def build_mermaid(results: list[dict], base_url: str,
                  navigations: dict[str, list[dict]] | None = None) -> str:
    """Build a Mermaid LR flowchart grouped by navigation sections.

    Each navigation (Main, Footer, Sidebar…) gets its own sub-tree
    so the diagram mirrors the site's actual IA.
    """
    url_to_page = _url_to_page(results)
    make_id = _make_mermaid_id_factory()

    lines = ["graph LR"]
    node_classes: list[str] = []

    unique_cats = sorted({p.get("category", CAT_FALLBACK) for p in results} | {"Home", CAT_FALLBACK})
    palette = category_palette_map(unique_cats)
    style_defs: list[str] = []
    cat_class: dict[str, str] = {}
    for i, cat in enumerate(unique_cats):
        cls = f"cat{i}"
        cat_class[cat] = cls
        color = palette[cat]
        lbl = _contrast_label_hex(color)
        style_defs.append(
            f"    classDef {cls} fill:{color},stroke:#64748b,stroke-width:1.2px,color:{lbl}"
        )

    root_id = make_id("ROOT")
    root_label = urlparse(base_url).netloc
    home_page = url_to_page.get(normalize_url(base_url))
    if home_page and home_page.get("title"):
        root_label = home_page["title"][:45]
    root_label = root_label.replace('"', "'")
    lines.append(f'    {root_id}["{root_label}"]')
    if home_page:
        node_classes.append(
            f"    class {root_id} {cat_class.get(home_page.get('category', CAT_FALLBACK), 'cat0')}"
        )

    rendered_urls: set[str] = set()
    if home_page:
        rendered_urls.add(home_page["url"])
    node_count = 0
    max_nodes = 100

    def _add_menu_node(item: dict, parent_id: str):
        nonlocal node_count
        if node_count >= max_nodes:
            return
        node_count += 1

        label = item["label"][:45].replace('"', "'")
        page = url_to_page.get(item["url"]) if item["url"] else None
        if page:
            rendered_urls.add(page["url"])
            if page.get("title"):
                label = page["title"][:45].replace('"', "'")
            cat = page.get("category", CAT_FALLBACK)
        else:
            cat = CAT_FALLBACK

        nid = make_id(item.get("label", "item"))
        lines.append(f'    {nid}["{label}"]')
        lines.append(f"    {parent_id} --> {nid}")
        node_classes.append(f"    class {nid} {cat_class.get(cat, 'cat0')}")

        for child in item.get("children", []):
            _add_menu_node(child, nid)

    if navigations:
        for nav_name, nav_items in navigations.items():
            section_id = make_id(nav_name)
            safe_name = nav_name.replace('"', "'")
            lines.append(f'    {section_id}["{safe_name}"]')
            lines.append(f"    {root_id} --> {section_id}")
            style_defs.append(
                f"    style {section_id} fill:#EDE9FE,stroke:#78716C,"
                f"stroke-width:2px,color:#111827,font-weight:bold"
            )
            for item in nav_items:
                _add_menu_node(item, section_id)

        remaining = [p for p in results if p["url"] not in rendered_urls
                     and p.get("status_code", 200) != 404]
        if remaining and node_count < max_nodes:
            other_id = make_id("altre_pagine")
            lines.append(f'    {other_id}["Altre pagine ({len(remaining)})"]')
            lines.append(f"    {root_id} --> {other_id}")
            node_classes.append(f"    class {other_id} {cat_class.get(CAT_FALLBACK, 'cat0')}")
            for page in remaining[:max_nodes - node_count]:
                node_count += 1
                nid = make_id(page["url"])
                plabel = (page.get("title") or page["url"])[:45].replace('"', "'")
                cat = page.get("category", CAT_FALLBACK)
                lines.append(f'    {nid}["{plabel}"]')
                lines.append(f"    {other_id} --> {nid}")
                node_classes.append(f"    class {nid} {cat_class.get(cat, 'cat0')}")
    else:
        tree, path_to_page = _build_page_tree(results)

        def walk(tree_node: dict, parent_id: str, current_path: str = ""):
            nonlocal node_count
            for key in sorted(tree_node.keys()):
                if node_count >= max_nodes:
                    return
                full_path = f"{current_path}/{key}".strip("/") if current_path else key
                nid = make_id(full_path)
                node_count += 1
                page = path_to_page.get(full_path)
                if page:
                    label = (page.get("title") or key)[:45].replace('"', "'")
                    cat = page.get("category", CAT_FALLBACK)
                else:
                    label = key
                    cat = CAT_FALLBACK
                lines.append(f'    {nid}["{label}"]')
                lines.append(f"    {parent_id} --> {nid}")
                node_classes.append(f"    class {nid} {cat_class.get(cat, 'cat0')}")
                walk(tree_node[key], nid, full_path)

        walk(tree, root_id)

    lines.extend(node_classes)
    lines.extend(style_defs)
    return "\n".join(lines)



def render_mermaid_html(mermaid_code: str, height: int = 600,
                        show_download: bool = False) -> str:
    """Return a self-contained HTML page that renders a Mermaid diagram (zoom + opz. download)."""
    download_btn_css = ""
    download_btn_html = ""
    download_btn_js = ""

    if show_download:
        download_btn_css = """
        #dl-btn {
            position: fixed; top: 12px; right: 16px; z-index: 100;
            background: #111827; color: #fff; border: none;
            padding: 8px 18px; border-radius: 8px; cursor: pointer;
            font-family: 'Inter', sans-serif; font-size: 13px; font-weight: 500;
            box-shadow: 0 2px 6px rgba(0,0,0,.15);
            transition: background .2s, filter .2s;
        }
        #dl-btn:hover { background: #27272a; filter: brightness(1.05); }
        """
        download_btn_html = '<button type="button" id="dl-btn">Scarica JPEG</button>'
        download_btn_js = """
        document.getElementById('dl-btn').addEventListener('click', function() {
            var svg = document.querySelector('#diagram-inner svg');
            if (!svg) return;
            var svgData = new XMLSerializer().serializeToString(svg);
            var canvas = document.createElement('canvas');
            var ctx = canvas.getContext('2d');
            var img = new Image();
            img.onload = function() {
                var sc = 2;
                canvas.width = img.width * sc;
                canvas.height = img.height * sc;
                ctx.fillStyle = '#ffffff';
                ctx.fillRect(0, 0, canvas.width, canvas.height);
                ctx.drawImage(img, 0, 0, canvas.width, canvas.height);
                var a = document.createElement('a');
                a.download = 'sitemap_diagram.jpg';
                a.href = canvas.toDataURL('image/jpeg', 0.95);
                a.click();
            };
            img.src = 'data:image/svg+xml;base64,' + btoa(unescape(encodeURIComponent(svgData)));
        });
        """

    return f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap" rel="stylesheet">
    <script src="https://cdn.jsdelivr.net/npm/mermaid@11/dist/mermaid.min.js"></script>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            padding: 16px 20px 24px;
            background: #f9fafb;
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
        }}
        #zoom-toolbar {{
            position: fixed; top: 12px; left: 50%; transform: translateX(-50%); z-index: 101;
            display: flex; align-items: center; gap: 6px;
            background: #fff; border: 1px solid #eaedf0; border-radius: 10px;
            padding: 6px 10px; box-shadow: 0 2px 8px rgba(0,0,0,.08);
        }}
        #zoom-toolbar button {{
            font-family: 'Inter', sans-serif; font-size: 13px; font-weight: 500;
            border: 1px solid #e5e7eb; background: #f3f4f6; color: #1f2937;
            padding: 6px 12px; border-radius: 8px; cursor: pointer;
            min-width: 40px;
        }}
        #zoom-toolbar button:hover {{ background: #e5e7eb; }}
        #zoom-toolbar #zoom-pct {{ min-width: 52px; text-align: center; font-size: 12px; color: #4b5563; }}
        #viewport {{
            margin-top: 52px;
            overflow: auto;
            max-height: calc(100vh - 72px);
            background: #eef0f3;
            border: 1px solid #eaedf0;
            border-radius: 12px;
            padding: 16px;
        }}
        #zoom-inner {{
            transform-origin: 0 0;
            transition: transform 0.12s ease-out;
            display: inline-block;
        }}
        #diagram-inner {{
            background: #fff;
            border-radius: 12px;
            padding: 28px 32px;
            border: 1px solid #eaedf0;
        }}
        .mermaid svg {{ height: auto; }}
        {download_btn_css}
    </style>
</head>
<body>
    <div id="zoom-toolbar">
        <button type="button" id="zoom-out" title="Zoom indietro">−</button>
        <span id="zoom-pct">100%</span>
        <button type="button" id="zoom-in" title="Zoom avanti">+</button>
        <button type="button" id="zoom-reset" title="Ripristina 100%">Reset</button>
    </div>
    {download_btn_html}
    <div id="viewport">
        <div id="zoom-inner">
            <div id="diagram-inner">
                <pre class="mermaid">
{mermaid_code}
                </pre>
            </div>
        </div>
    </div>
    <script>
        (function() {{
            var scale = 1;
            var inner = document.getElementById('zoom-inner');
            var pct = document.getElementById('zoom-pct');
            var vp = document.getElementById('viewport');
            function apply() {{
                inner.style.transform = 'scale(' + scale + ')';
                pct.textContent = Math.round(scale * 100) + '%';
            }}
            document.getElementById('zoom-in').addEventListener('click', function() {{
                scale = Math.min(scale + 0.2, 4);
                apply();
            }});
            document.getElementById('zoom-out').addEventListener('click', function() {{
                scale = Math.max(scale - 0.2, 0.25);
                apply();
            }});
            document.getElementById('zoom-reset').addEventListener('click', function() {{
                scale = 1;
                apply();
            }});
            vp.addEventListener('wheel', function(e) {{
                if (!e.ctrlKey) return;
                e.preventDefault();
                var d = e.deltaY < 0 ? 1.08 : 0.92;
                scale = Math.min(Math.max(scale * d, 0.25), 4);
                apply();
            }}, {{ passive: false }});
            apply();
        }})();
        mermaid.initialize({{
            startOnLoad: true,
            theme: 'neutral',
            themeVariables: {{
                primaryColor: '#f1f5f9',
                primaryTextColor: '#111827',
                primaryBorderColor: '#94a3b8',
                lineColor: '#64748b',
                secondaryColor: '#e2e8f0',
                tertiaryColor: '#f8fafc',
            }},
            flowchart: {{ useMaxWidth: false, htmlLabels: true, curve: 'basis' }},
            securityLevel: 'loose'
        }});
        {download_btn_js}
    </script>
</body>
</html>"""



# ─────────────────────────────────────────────
# Excel export & share packs
# ─────────────────────────────────────────────

def _excel_safe_str(val, max_len: int = 32700) -> str:
    """Strip characters illegal in Excel/OpenXML cells."""
    if val is None:
        return ""
    s = str(val)
    s = _EXCEL_CTRL_RE.sub(" ", s)
    return s if len(s) <= max_len else s[:max_len]


def build_share_pack(start_url: str, results: list[dict],
                     navigations: dict[str, list[dict]], errors_404: list[str]) -> dict:
    """Serializable snapshot for URL or file sharing (_html omitted)."""
    clean_results = []
    for p in results:
        d = {k: v for k, v in p.items() if k != "_html"}
        clean_results.append(d)
    return {
        "v": SHARE_PACK_VERSION,
        "start_url": start_url,
        "results": clean_results,
        "navigations": navigations,
        "errors_404": list(errors_404),
    }


def share_pack_to_json_bytes(pack: dict) -> bytes:
    return json.dumps(pack, ensure_ascii=False, indent=2).encode("utf-8")


def encode_share_query_payload(pack: dict) -> str:
    raw = json.dumps(pack, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    z = zlib.compress(raw, level=9)
    return base64.urlsafe_b64encode(z).decode("ascii").rstrip("=")


def decode_share_query_payload(token: str) -> dict:
    s = token.strip()
    pad = (-len(s)) % 4
    if pad:
        s += "=" * pad
    z = base64.urlsafe_b64decode(s.encode("ascii"))
    raw = zlib.decompress(z)
    data = json.loads(raw.decode("utf-8"))
    if data.get("v") != SHARE_PACK_VERSION:
        raise ValueError("Unsupported share pack version")
    return data


def apply_share_pack(data: dict) -> None:
    """Restore session from a share pack dict."""
    start_url = data["start_url"]
    results = data["results"]
    navigations = data.get("navigations") or {}
    errors_404 = data.get("errors_404") or []
    for page in results:
        page.setdefault("_html", "")
        page["category"] = infer_page_category(page, start_url, navigations)
    st.session_state.results = results
    st.session_state.errors_404 = errors_404
    st.session_state.navigations = navigations
    st.session_state.crawl_start_url = start_url
    st.session_state.mermaid_code = build_mermaid(results, start_url, navigations)


def _excel_lookup_page(url: str, url_to_page: dict[str, dict]) -> dict | None:
    if not url:
        return None
    u = normalize_url(url)
    return url_to_page.get(u) or url_to_page.get(url)


def _ia_rows_from_navigations(
    navigations: dict[str, list[dict]], url_to_page: dict[str, dict],
) -> list[dict]:
    """One row per voce di menu in ordine gerarchico."""
    rows: list[dict] = []

    def walk(nav_zone: str, items: list[dict], parent_labels: list[str], level: int) -> None:
        for item in items:
            lab = item.get("label") or ""
            labels = parent_labels + [lab]
            path_str = " > ".join(labels)
            url = (item.get("url") or "").strip()
            page = _excel_lookup_page(url, url_to_page)
            title = (page.get("title") or "") if page else ""
            cat = (page.get("category", CAT_FALLBACK) if page else CAT_FALLBACK)
            sc = int(page.get("status_code") or 0) if page else 0
            meta = (page.get("meta_description") or "") if page else ""
            h1 = (page.get("h1") or "") if page else ""
            h2s = page.get("h2_list") or [] if page else []
            if not isinstance(h2s, list):
                h2s = [str(h2s)]
            wc = int(page.get("word_count") or 0) if page else 0
            bc = (page.get("breadcrumbs") or "") if page else ""
            dep = int(page.get("depth") or 0) if page else 0
            rows.append({
                "zone": nav_zone,
                "level": level,
                "path": path_str,
                "menu_label": lab,
                "url": url,
                "title": title,
                "category": cat,
                "status": sc,
                "meta_description": meta,
                "h1": h1,
                "h2": "; ".join(str(x) for x in h2s),
                "word_count": wc,
                "breadcrumbs": bc,
                "depth": dep,
            })
            ch = item.get("children") or []
            if ch:
                walk(nav_zone, ch, labels, level + 1)

    for zone, top_items in navigations.items():
        walk(zone, top_items, [], 0)
    return rows


def _ia_rows_from_url_tree(results: list[dict]) -> list[dict]:
    """Fallback: gerarchia come nel tab Sitemap senza menu."""

    def _url_path_key(u: str) -> str:
        pp = urlparse(u)
        pparts = [x for x in pp.path.strip("/").split("/") if x]
        return "/".join(pparts) if pparts else "(home)"

    path_to_title: dict[str, str] = {}
    path_to_page: dict[str, dict] = {}
    for page in results:
        fk = _url_path_key(page["url"])
        path_to_page[fk] = page
        title = (page.get("title") or "").strip()
        if title:
            path_to_title[fk] = title

    tree: dict = {}
    for page in results:
        fk = _url_path_key(page["url"])
        parts = ["(home)"] if fk == "(home)" else fk.split("/")
        node = tree
        for part in parts:
            if part not in node:
                node[part] = {}
            node = node[part]

    rows: list[dict] = []

    def walk_tree(node: dict, parent_labels: list[str], level: int, current_path: str) -> None:
        for key in sorted(node.keys()):
            full_path = f"{current_path}/{key}".strip("/") if current_path else key
            label = path_to_title.get(full_path, key)
            matched = path_to_page.get(full_path)
            h2s = matched.get("h2_list") or [] if matched else []
            if matched and not isinstance(h2s, list):
                h2s = [str(h2s)]
            path_str = " > ".join(parent_labels + [label])
            rows.append({
                "zone": "Struttura URL (fallback)",
                "level": level,
                "path": path_str,
                "menu_label": label,
                "url": matched["url"] if matched else "",
                "title": (matched.get("title") or "") if matched else "",
                "category": (matched.get("category", CAT_FALLBACK) if matched else CAT_FALLBACK),
                "status": int(matched.get("status_code") or 0) if matched else 0,
                "meta_description": (matched.get("meta_description") or "") if matched else "",
                "h1": (matched.get("h1") or "") if matched else "",
                "h2": "; ".join(str(x) for x in h2s) if matched else "",
                "word_count": int(matched.get("word_count") or 0) if matched else 0,
                "breadcrumbs": (matched.get("breadcrumbs") or "") if matched else "",
                "depth": int(matched.get("depth") or 0) if matched else 0,
            })
            walk_tree(node[key], parent_labels + [label], level + 1, full_path)

    walk_tree(tree, [], 0, "")
    return rows


def generate_excel(
    results: list[dict],
    navigations: dict[str, list[dict]] | None = None,
    mermaid_code: str = "",
) -> bytes:
    """Workbook: IA, Sitemap come in app, immagine diagramma Mermaid."""
    wb = openpyxl.Workbook()
    url_to_page = {p["url"]: p for p in results}

    headers = [
        "Zona navigazione", "Livello", "Percorso (IA)", "Voce menu",
        "URL", "Title pagina", "Categoria", "Status",
        "Meta description", "H1", "H2 (lista)", "Word count", "Breadcrumbs", "Depth crawl",
    ]
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(bottom=Side(style="thin", color="DDDDDD"))

    ws1 = wb.active
    ws1.title = "Information Architecture"

    for col_idx, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ia_rows: list[dict] = []
    if navigations:
        ia_rows = _ia_rows_from_navigations(navigations, url_to_page)
        in_nav = set(flatten_nav_urls(navigations))
        home_guess = normalize_url(
            results[0]["url"].split(urlparse(results[0]["url"]).path)[0] + "/"
        ) if results else ""
        for page in sorted(results, key=lambda p: p["url"]):
            u = page["url"]
            if u in in_nav or page.get("status_code") == 404:
                continue
            if home_guess and u == home_guess:
                continue
            h2s = page.get("h2_list") or []
            if not isinstance(h2s, list):
                h2s = [str(h2s)]
            ia_rows.append({
                "zone": "Altre pagine (non in menu)",
                "level": 0,
                "path": page.get("title") or u,
                "menu_label": page.get("title") or urlparse(u).path or u,
                "url": u,
                "title": page.get("title") or "",
                "category": page.get("category", CAT_FALLBACK),
                "status": int(page.get("status_code") or 0),
                "meta_description": page.get("meta_description") or "",
                "h1": page.get("h1") or "",
                "h2": "; ".join(str(x) for x in h2s),
                "word_count": int(page.get("word_count") or 0),
                "breadcrumbs": page.get("breadcrumbs") or "",
                "depth": int(page.get("depth") or 0),
            })
    else:
        ia_rows = _ia_rows_from_url_tree(results)

    for row_idx, row in enumerate(ia_rows, 2):
        ws1.cell(row=row_idx, column=1, value=_excel_safe_str(row["zone"], max_len=200))
        ws1.cell(row=row_idx, column=2, value=int(row["level"]))
        ws1.cell(row=row_idx, column=3, value=_excel_safe_str(row["path"]))
        ws1.cell(row=row_idx, column=4, value=_excel_safe_str(row["menu_label"]))
        ws1.cell(row=row_idx, column=5, value=_excel_safe_str(row["url"]))
        ws1.cell(row=row_idx, column=6, value=_excel_safe_str(row["title"]))
        ws1.cell(row=row_idx, column=7, value=_excel_safe_str(row["category"]))
        sc = int(row["status"])
        status_cell = ws1.cell(row=row_idx, column=8, value=sc)
        if sc == 404:
            status_cell.font = Font(color="CC0000", bold=True)
        elif sc >= 400:
            status_cell.font = Font(color="FF6600", bold=True)
        else:
            status_cell.font = Font(color="228B22")
        ws1.cell(row=row_idx, column=9, value=_excel_safe_str(row["meta_description"]))
        ws1.cell(row=row_idx, column=10, value=_excel_safe_str(row["h1"]))
        ws1.cell(row=row_idx, column=11, value=_excel_safe_str(row["h2"]))
        ws1.cell(row=row_idx, column=12, value=int(row["word_count"]))
        ws1.cell(row=row_idx, column=13, value=_excel_safe_str(row["breadcrumbs"]))
        ws1.cell(row=row_idx, column=14, value=int(row["depth"]))
        for c in range(1, len(headers) + 1):
            ws1.cell(row=row_idx, column=c).border = thin_border

    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            (len(str(ws1.cell(row=r, column=col_idx).value or ""))
             for r in range(1, min(len(ia_rows) + 2, 80))),
            default=10,
        )
        ws1.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 55)

    # ── Sheet: Sitemap (livelli + anteprima ASCII + hyperlink URL) ──
    ws_map = wb.create_sheet("Sitemap")
    sm_rows = build_sitemap_tree_excel_rows(results, navigations)
    max_depth = 1
    if sm_rows:
        max_depth = min(
            max(len(r.get("levels") or []) for r in sm_rows),
            SM_EXCEL_MAX_LEVELS,
        )
    level_headers = [f"Livello {i}" for i in range(1, max_depth + 1)]
    sm_headers = (
        ["Navigazione / zona"]
        + level_headers
        + ["Anteprima albero (ASCII)", "URL (clic per aprire)"]
    )
    n_h = len(sm_headers)
    for col_idx, h in enumerate(sm_headers, 1):
        cell = ws_map.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    preview_col = 2 + max_depth
    url_col = preview_col + 1

    for row_idx, row in enumerate(sm_rows, 2):
        ws_map.cell(row=row_idx, column=1, value=_excel_safe_str(row["navigazione"], max_len=200))
        lv = (row.get("levels") or [])[:SM_EXCEL_MAX_LEVELS]
        for li in range(max_depth):
            val = _excel_safe_str(lv[li], max_len=500) if li < len(lv) else ""
            ws_map.cell(row=row_idx, column=2 + li, value=val)
        ws_map.cell(row=row_idx, column=preview_col, value=_excel_safe_str(row.get("albero", "")))
        u_raw = (row.get("url") or "").strip()
        u_cell = ws_map.cell(row=row_idx, column=url_col)
        if u_raw.startswith(("http://", "https://")):
            u_cell.value = _excel_url_display(u_raw)
            u_cell.hyperlink = u_raw
            u_cell.font = Font(color="0563C1", underline="single")
        else:
            u_cell.value = _excel_safe_str(u_raw, max_len=2048)
        for c in range(1, n_h + 1):
            ws_map.cell(row=row_idx, column=c).border = thin_border

    ws_map.column_dimensions["A"].width = 28
    for li in range(max_depth):
        ws_map.column_dimensions[get_column_letter(2 + li)].width = 26
    ws_map.column_dimensions[get_column_letter(preview_col)].width = 52
    ws_map.column_dimensions[get_column_letter(url_col)].width = 44

    last_row = len(sm_rows) + 1
    if last_row >= 1 and n_h >= 1:
        ws_map.auto_filter.ref = f"A1:{get_column_letter(n_h)}{last_row}"
        ws_map.freeze_panes = f"B2"

    hint_r = last_row + 2
    ws_map.merge_cells(
        start_row=hint_r, start_column=1, end_row=hint_r, end_column=max(n_h, 1),
    )
    hc = ws_map.cell(row=hint_r, column=1, value=(
        "Suggerimento: filtri sulla riga 1; congela già attivo (colonna zona). "
        "La colonna URL contiene collegamenti cliccabili al sito."
    ))
    hc.font = Font(size=9, italic=True, color="6B7280")
    hc.alignment = Alignment(wrap_text=True, vertical="top")

    # ── Sheet: Diagramma (PNG del Mermaid) ──
    ws_diag = wb.create_sheet("Diagramma")
    ws_diag.merge_cells("A1:F1")
    c1 = ws_diag.cell(row=1, column=1, value="Diagramma gerarchico (equivalente al tab Diagramma nell'app).")
    c1.font = Font(bold=True, size=12)
    png_bytes = mermaid_to_png_bytes(mermaid_code)
    if png_bytes:
        try:
            img = XLImage(io.BytesIO(png_bytes))
            max_w = 920
            if img.width > max_w:
                ratio = max_w / float(img.width)
                img.width = int(img.width * ratio)
                img.height = int(img.height * ratio)
            ws_diag.add_image(img, "A3")
            # Altezza riga in punti (~px * 0.75)
            ws_diag.row_dimensions[3].height = min(max(img.height * 0.75, 180), 380)
        except Exception:
            ws_diag.cell(row=3, column=1, value="Impossibile inserire l'immagine nel foglio.")
    else:
        ws_diag.cell(
            row=3, column=1,
            value=(
                "Immagine non generata: codice Mermaid vuoto, rete non disponibile o servizio di rendering "
                "non raggiungibile. Puoi usare il tab Diagramma nell'app o il file .md Mermaid in Esporta."
            ),
        )
        ws_diag["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_diag.column_dimensions["A"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═════════════════════════════════════════════
# SIDEBAR
# ═════════════════════════════════════════════
with st.sidebar:
    st.markdown(
        '<p class="sidebar-app-title">UX ARCHITECT PRO</p>',
        unsafe_allow_html=True,
    )
    st.markdown("### Fonte dati")
    st.markdown(
        '<p class="sidebar-source-hint">Inserisci un URL per un nuovo crawl oppure '
        "carica un rapporto <strong>.json</strong> esportato dalla scheda Esporta.</p>",
        unsafe_allow_html=True,
    )

    _source = st.radio(
        "Modalita",
        options=["url", "json"],
        format_func=lambda x: (
            "Nuovo crawl da URL" if x == "url" else "Rapporto condiviso (.json)"
        ),
        key="sidebar_data_source",
        label_visibility="collapsed",
    )

    start_url = ""
    max_depth = 3
    max_pages = 100
    run_crawl = False

    if _source == "url":
        st.markdown('<p class="sidebar-panel-label">URL del sito</p>', unsafe_allow_html=True)
        start_url = st.text_input(
            "URL di partenza",
            placeholder="https://www.example.com",
            help="URL completo da cui parte la scansione",
            label_visibility="collapsed",
            key="sidebar_url_input",
        )
        max_depth = st.slider(
            "Profondita massima", 1, 10, 3,
            help="Profondità massima di navigazione",
            key="sidebar_max_depth",
        )
        max_pages = st.slider(
            "Pagine massime", 10, 500, 100, step=10,
            help="Numero massimo di pagine da scansionare",
            key="sidebar_max_pages",
        )
        st.markdown("---")
        run_crawl = st.button(
            "Avvia Crawl", type="primary", use_container_width=True, key="btn_run_crawl",
        )
    else:
        st.markdown('<p class="sidebar-panel-label">File rapporto</p>', unsafe_allow_html=True)
        st.caption(
            "Formato esportato da **Esporta → Scarica rapporto condivisibile (.json)**"
        )
        _up_share = st.file_uploader(
            "Carica file JSON",
            type=["json"],
            help="Ricostruisce Sitemap, Diagramma, Tabella e export senza rifare il crawl",
            label_visibility="collapsed",
            key="sidebar_json_upload",
        )
        _import_clicked = st.button(
            "Carica rapporto",
            type="primary",
            use_container_width=True,
            disabled=_up_share is None,
            key="btn_import_share_json",
        )
        if _import_clicked and _up_share is not None:
            try:
                _raw = _up_share.getvalue()
                _data = json.loads(_raw.decode("utf-8"))
                if _data.get("v") != SHARE_PACK_VERSION:
                    st.error("Versione del file non supportata.")
                else:
                    apply_share_pack(_data)
                    st.session_state._loaded_share_token = None
                    if "r" in st.query_params:
                        del st.query_params["r"]
                    st.success("Rapporto caricato.")
                    st.rerun()
            except Exception as _exc:
                st.error(f"File non valido: {_exc}")

    st.markdown("---")
    st.markdown("### Legenda categorie")
    _res = st.session_state.get("results")
    if _res:
        _leg_pal = category_palette_map([p.get("category", CAT_FALLBACK) for p in _res])
        for _cat in sorted(_leg_pal.keys()):
            _col = _leg_pal[_cat]
            st.markdown(
                f'<div class="legend-item">'
                f'<span class="legend-dot" style="background:{_col}"></span>'
                f'{_cat}</div>',
                unsafe_allow_html=True,
            )
    else:
        st.caption(
            "Le categorie compaiono qui dopo un crawl o l’import di un rapporto: "
            "derivano da menu e percorsi URL del sito."
        )


# ═════════════════════════════════════════════
# MAIN AREA
# ═════════════════════════════════════════════

if "results" not in st.session_state:
    st.session_state.results = None
    st.session_state.errors_404 = []
    st.session_state.navigations = {}
    st.session_state.mermaid_code = ""
    st.session_state.crawl_start_url = ""
    st.session_state._loaded_share_token = None

# Apri rapporto da link (?r=... compresso)
_r_raw = st.query_params.get("r")
_r_q = _r_raw[0] if isinstance(_r_raw, list) and _r_raw else _r_raw
if _r_q and _r_q != st.session_state.get("_loaded_share_token"):
    try:
        _pack = decode_share_query_payload(_r_q)
        apply_share_pack(_pack)
        st.session_state._loaded_share_token = _r_q
    except Exception:
        st.session_state._loaded_share_token = _r_q
        st.error(
            "Il parametro **r** nell'URL non e valido (link troncato o versione non compatibile). "
            "Chiedi al collega il file **.json** oppure un nuovo link."
        )

if run_crawl:
    if not start_url:
        st.error("Inserisci un URL valido per avviare il crawl.")
    else:
        if not start_url.startswith(("http://", "https://")):
            start_url = "https://" + start_url

        st.markdown('<div class="section-title">Crawl in corso</div>', unsafe_allow_html=True)
        progress_bar = st.progress(0, text="Avvio crawl…")
        status_text = st.empty()
        log_container = st.empty()

        results, errors_404, navigations = crawl_site(
            start_url, max_depth, max_pages,
            progress_bar, log_container, status_text,
        )

        for page in results:
            page["category"] = infer_page_category(page, start_url, navigations)

        st.session_state.results = results
        st.session_state.errors_404 = errors_404
        st.session_state.navigations = navigations
        st.session_state.crawl_start_url = start_url
        st.session_state.mermaid_code = build_mermaid(results, start_url, navigations)
        st.session_state._loaded_share_token = None
        if "r" in st.query_params:
            del st.query_params["r"]
        st.rerun()


# ─────────────────────────────────────────────
# Results dashboard
# ─────────────────────────────────────────────
if st.session_state.results is not None:
    results = st.session_state.results
    errors_404 = st.session_state.errors_404
    _cat_colors_ui = category_palette_map([p.get("category", CAT_FALLBACK) for p in results])

    # KPI cards
    total_pages = len(results)
    total_words = sum(p.get("word_count", 0) for p in results)
    avg_words = total_words // max(total_pages, 1)
    n_categories = len(set(p.get("category", CAT_FALLBACK) for p in results))
    n_404 = len(errors_404)

    st.markdown('<div class="section-title">Riepilogo</div>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f'<div class="stat-card"><h3>Pagine totali</h3>'
                     f'<div class="value">{total_pages}</div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="stat-card"><h3>Parole (media)</h3>'
                     f'<div class="value">{avg_words}</div></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="stat-card"><h3>Categorie</h3>'
                     f'<div class="value">{n_categories}</div></div>', unsafe_allow_html=True)
    with c4:
        color_404 = "#b91c1c" if n_404 > 0 else "#047857"
        st.markdown(f'<div class="stat-card"><h3>Errori 404</h3>'
                     f'<div class="value" style="color:{color_404}">{n_404}</div></div>', unsafe_allow_html=True)

    st.markdown("---")

    # ── Tabs (Sitemap e Diagramma per primi) ──
    tab_sitemap, tab_mermaid, tab_table, tab_export = st.tabs(
        ["Sitemap", "Diagramma", "Tabella", "Esporta"]
    )

    with tab_sitemap:
        st.markdown('<div class="section-title">Information Architecture</div>', unsafe_allow_html=True)

        navigations = st.session_state.navigations
        url_to_page = {p["url"]: p for p in results}

        def render_menu_tree(items: list[dict], prefix: str = ""):
            lines_out: list[str] = []
            for i, item in enumerate(items):
                last = i == len(items) - 1
                connector = "└── " if last else "├── "
                page = url_to_page.get(item["url"]) if item["url"] else None
                label = item["label"]
                if page and page.get("title"):
                    label = page["title"]
                lines_out.append(f"{prefix}{connector}{label}")
                extension = "    " if last else "│   "
                if item.get("children"):
                    lines_out.extend(render_menu_tree(item["children"], prefix + extension))
            return lines_out

        if navigations:
            st.markdown(
                '<div class="section-subtitle">'
                'Struttura delle navigazioni estratte dal sito'
                '</div>',
                unsafe_allow_html=True,
            )

            home_page = url_to_page.get(normalize_url(
                results[0]["url"].split(urlparse(results[0]["url"]).path)[0] + "/"
            ))
            root_title = (home_page.get("title") if home_page else None) or urlparse(results[0]["url"]).netloc

            for nav_name, nav_items in navigations.items():
                tree_text = f"{nav_name}\n"
                tree_text += "\n".join(render_menu_tree(nav_items))
                st.code(tree_text, language=None)

            all_nav_urls = set(flatten_nav_urls(navigations))
            home_url = normalize_url(
                results[0]["url"].split(urlparse(results[0]["url"]).path)[0] + "/"
            )
            remaining = [p for p in results
                         if p["url"] not in all_nav_urls
                         and p.get("status_code", 200) != 404
                         and p["url"] != home_url]
            if remaining:
                tree_text = f"Altre pagine ({len(remaining)})\n"
                tree_lines: list[str] = []
                for j, page in enumerate(remaining[:30]):
                    last = j == min(len(remaining), 30) - 1
                    conn = "└── " if last else "├── "
                    tree_lines.append(f"{conn}{page.get('title') or page['url']}")
                tree_text += "\n".join(tree_lines)
                st.code(tree_text, language=None)

        else:
            st.markdown(
                '<div class="section-subtitle">'
                'Struttura gerarchica basata sui percorsi URL'
                '</div>',
                unsafe_allow_html=True,
            )

            path_to_title: dict[str, str] = {}
            for page in results:
                parsed = urlparse(page["url"])
                path = parsed.path.strip("/") or "(home)"
                title = (page.get("title") or "").strip()
                if title:
                    path_to_title[path] = title

            tree: dict = {}
            for page in results:
                parsed = urlparse(page["url"])
                parts = [p for p in parsed.path.strip("/").split("/") if p]
                if not parts:
                    parts = ["(home)"]
                node = tree
                for part in parts:
                    if part not in node:
                        node[part] = {}
                    node = node[part]

            def render_tree(node: dict, prefix: str = "", is_last: bool = True,
                            depth: int = 0, current_path: str = ""):
                lines_out: list[str] = []
                items = sorted(node.keys())
                for i, key in enumerate(items):
                    last = i == len(items) - 1
                    connector = "└── " if last else "├── "
                    full_path = f"{current_path}/{key}".strip("/") if current_path else key
                    label = path_to_title.get(full_path, key)
                    lines_out.append(f"{prefix}{connector}{label}")
                    extension = "    " if last else "│   "
                    lines_out.extend(render_tree(
                        node[key], prefix + extension, last, depth + 1, full_path
                    ))
                return lines_out

            root_label = urlparse(results[0]["url"]).netloc
            root_title = path_to_title.get("(home)", root_label)
            tree_text = f"{root_title}\n"
            tree_text += "\n".join(render_tree(tree))
            st.code(tree_text, language=None)

    with tab_mermaid:
        st.markdown('<div class="section-title">Diagramma gerarchico</div>', unsafe_allow_html=True)
        st.markdown(
            '<div class="section-subtitle">Struttura del sito con layout orizzontale. '
            "Usa <strong>+ / − / Reset</strong> in alto al centro per lo zoom, oppure "
            "<strong>Ctrl + rotellina</strong> sul diagramma. Scroll nel riquadro per spostarti. "
            "Il pulsante in alto a destra scarica il JPEG.</div>",
            unsafe_allow_html=True,
        )

        mermaid_code = st.session_state.mermaid_code
        diagram_html = render_mermaid_html(mermaid_code, height=650, show_download=True)
        components.html(diagram_html, height=720, scrolling=True)

        with st.expander("Mostra codice Mermaid"):
            st.code(mermaid_code, language="mermaid")

    with tab_table:
        st.markdown('<div class="section-title">Pagine analizzate</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-subtitle">Filtra per categoria e clicca per espandere i dettagli</div>', unsafe_allow_html=True)

        filter_cat = st.multiselect(
            "Filtra per categoria",
            options=sorted(set(p.get("category", CAT_FALLBACK) for p in results)),
            default=sorted(set(p.get("category", CAT_FALLBACK) for p in results)),
        )
        filtered = [p for p in results if p.get("category", CAT_FALLBACK) in filter_cat]

        for page in filtered:
            status = page["status_code"]
            cat = page.get("category", CAT_FALLBACK)
            cat_color = _cat_colors_ui.get(cat, CATEGORY_COLOR_FALLBACK)

            if status == 404:
                status_label = "404"
            elif status >= 400:
                status_label = str(status)
            else:
                status_label = str(status)

            title_display = page.get("title") or page["url"][:60]
            with st.expander(f"`{status_label}` — **{title_display}**"):
                col_a, col_b = st.columns([2, 1])
                with col_a:
                    st.markdown(f"**URL:** `{page['url']}`")
                    st.markdown(f"**Title:** {page.get('title', '—')}")
                    st.markdown(f"**H1:** {page.get('h1', '—')}")
                    st.markdown(f"**Meta Description:** {page.get('meta_description', '—') or '—'}")
                    if page.get("h2_list"):
                        st.markdown("**H2:**")
                        for h2 in page["h2_list"]:
                            st.markdown(f"  - {h2}")
                with col_b:
                    st.markdown(
                        f'<span class="category-badge" style="background:{cat_color};'
                        f'color:{_contrast_label_hex(cat_color)}">{cat}</span>',
                        unsafe_allow_html=True,
                    )
                    st.metric("Word Count", page.get("word_count", 0))
                    st.metric("Depth", page.get("depth", 0))
                    if page.get("breadcrumbs"):
                        st.markdown(f"**Breadcrumbs:** {page['breadcrumbs']}")

    with tab_export:
        st.markdown('<div class="section-title">Esportazione dati</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-subtitle">Scarica i risultati nei formati disponibili</div>', unsafe_allow_html=True)

        _start_share = (st.session_state.get("crawl_start_url") or "").strip()
        if not _start_share and results:
            _p0 = urlparse(results[0]["url"])
            _start_share = f"{_p0.scheme}://{_p0.netloc}/"
        _share_pack = build_share_pack(
            _start_share, results, st.session_state.navigations, errors_404,
        )
        _share_json = share_pack_to_json_bytes(_share_pack)
        try:
            _share_enc = encode_share_query_payload(_share_pack)
        except Exception:
            _share_enc = ""

        st.markdown("##### Condividi il rapporto")
        st.caption(
            "Altri possono vedere lo stesso risultato aprendo un link oppure importando il file .json "
            "(sidebar → Rapporto condiviso)."
        )
        st.download_button(
            "Scarica rapporto condivisibile (.json)",
            data=_share_json,
            file_name="ux_architect_pro_rapporto.json",
            mime="application/json",
            use_container_width=True,
            key="dl_share_json",
        )
        if _share_enc and len(_share_enc) <= SHARE_URL_MAX_CHARS:
            if st.button("Aggiorna URL con link di condivisione", use_container_width=True, key="btn_set_share_url"):
                st.query_params["r"] = _share_enc
            st.caption(
                "Dopo il clic, copia l'indirizzo completo dalla barra del browser e invialo: aprendolo si carica questo rapporto."
            )
        elif _share_enc:
            st.info(
                f"Questo rapporto e troppo grande per un link nell'URL (~{len(_share_enc)} caratteri). "
                "Usa il file **.json** sopra."
            )

        st.markdown("---")

        col_xl, col_mmd = st.columns(2)

        with col_xl:
            st.markdown("**Excel — Information Architecture**")
            st.caption(
                "Fogli: **Information Architecture**, **Sitemap** (livelli gerarchici, anteprima ASCII, "
                "URL cliccabili, filtri), **Diagramma** (PNG)."
            )
            try:
                _excel_bytes = generate_excel(
                    results, st.session_state.navigations, st.session_state.mermaid_code,
                )
            except Exception as _xlsx_exc:
                _excel_bytes = b""
                st.error(f"Errore durante la creazione del file Excel: {_xlsx_exc}")
            if _excel_bytes:
                st.download_button(
                    "Scarica Excel",
                    data=_excel_bytes,
                    file_name="ux_architect_pro_information_architecture.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_excel",
                )

        with col_mmd:
            st.markdown("**Mermaid.js diagram**")
            st.caption("Codice Mermaid pronto per qualsiasi renderer")
            st.download_button(
                "Scarica Mermaid",
                data=f"```mermaid\n{st.session_state.mermaid_code}\n```",
                file_name="sitemap_mermaid.md",
                mime="text/markdown",
                use_container_width=True,
                key="dl_mermaid",
            )

        if errors_404:
            st.markdown("---")
            st.markdown("**Errori 404**")
            for u in errors_404:
                st.markdown(f"- `{u}`")

else:
    st.markdown("""
<div class="ux-empty-hero">
  <div class="ux-empty-stack">
  <h1 class="ux-empty-title">Ciao! <span class="ux-wave-hand" role="img" aria-label="mano che saluta">👋</span></h1>
  <h3 class="ux-empty-subtitle">Mappiamo insieme</h3>
  <p class="ux-empty-hint">
    Inserisci un URL per un nuovo crawl oppure carica un rapporto .json che ti ha girato un tuo collega. 😊
  </p>
  </div>
</div>
    """, unsafe_allow_html=True)
